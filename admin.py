"""Admin-Bereich: gesicherte Ansicht aller eingetragenen E-Mails + Excel-Export.

Sicherheitskonzept
------------------
* Keine Standard-Zugangsdaten: Ohne gesetzte Umgebungsvariablen ist der Login
  komplett gesperrt (kein "admin/admin" ab Werk).
* Passwortpruefung in Konstantzeit (hmac.compare_digest bzw. Werkzeug-Hash),
  damit keine Timing-Angriffe moeglich sind.
* Session-basierter Login mit HttpOnly-/SameSite-/(prod) Secure-Cookies.
* CSRF-Token fuer alle veraendernden POST-Requests (Login/Logout).
* Einfache Brute-Force-Sperre pro IP.
* Excel-Export mit Schutz gegen Formel-Injection (CSV/Excel Injection).

Konfiguration ueber Umgebungsvariablen:
    ADMIN_USERNAME       Benutzername (Default: "admin")
    ADMIN_PASSWORD_HASH  Werkzeug-Passwort-Hash (bevorzugt)
    ADMIN_PASSWORD       Klartext-Passwort (Fallback, weniger sicher)
    FLASK_SECRET_KEY     Schluessel fuer signierte Sessions (in Prod PFLICHT)

Passwort-Hash erzeugen:
    python -c "from werkzeug.security import generate_password_hash as g; \
print(g(input('Passwort: ')))"
"""

import hmac
import os
import secrets
import time
from datetime import datetime
from functools import wraps
from io import BytesIO

from flask import (
    Blueprint,
    Response,
    abort,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from werkzeug.security import check_password_hash

import db

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")

# --- Zugangsdaten aus der Umgebung -----------------------------------------
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD_HASH = os.environ.get("ADMIN_PASSWORD_HASH")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD")  # Fallback (Klartext)

# --- Brute-Force-Schutz (pro IP, im Prozessspeicher) ------------------------
_MAX_ATTEMPTS = 5
_LOCKOUT_SECONDS = 300
_attempts = {}  # ip -> (fehlversuche, gesperrt_bis_ts)


def admin_configured():
    """True, wenn ueberhaupt ein Passwort hinterlegt ist."""
    return bool(ADMIN_PASSWORD_HASH or ADMIN_PASSWORD)


def _check_credentials(username, password):
    """Vergleich in Konstantzeit; beide Faktoren werden immer geprueft."""
    user_ok = hmac.compare_digest((username or ""), ADMIN_USERNAME)
    if ADMIN_PASSWORD_HASH:
        pass_ok = check_password_hash(ADMIN_PASSWORD_HASH, password or "")
    elif ADMIN_PASSWORD:
        pass_ok = hmac.compare_digest((password or ""), ADMIN_PASSWORD)
    else:
        pass_ok = False
    return user_ok and pass_ok


def _client_ip():
    # Hinter einem Reverse-Proxy liefert X-Forwarded-For die echte IP; wir
    # nehmen nur den ersten Eintrag. Ohne Proxy bleibt remote_addr.
    fwd = request.headers.get("X-Forwarded-For", "")
    if fwd:
        return fwd.split(",")[0].strip()
    return request.remote_addr or "unbekannt"


def _is_locked(ip):
    count, until = _attempts.get(ip, (0, 0))
    return time.time() < until


def _register_failure(ip):
    count, _ = _attempts.get(ip, (0, 0))
    count += 1
    until = time.time() + _LOCKOUT_SECONDS if count >= _MAX_ATTEMPTS else 0
    _attempts[ip] = (count, until)


def _reset_failures(ip):
    _attempts.pop(ip, None)


# --- CSRF -------------------------------------------------------------------
def csrf_token():
    tok = session.get("_csrf")
    if not tok:
        tok = secrets.token_urlsafe(32)
        session["_csrf"] = tok
    return tok


def _check_csrf():
    sent = request.form.get("_csrf", "")
    stored = session.get("_csrf", "")
    return bool(stored) and hmac.compare_digest(sent, stored)


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("admin_authenticated"):
            return redirect(url_for("admin.login", next=request.path))
        return view(*args, **kwargs)

    return wrapped


def _safe_next(target):
    """Open-Redirect verhindern: nur lokale, absolute Pfade zulassen."""
    if target and target.startswith("/") and not target.startswith("//"):
        return target
    return url_for("admin.dashboard")


@admin_bp.route("/login", methods=["GET", "POST"])
def login():
    if session.get("admin_authenticated"):
        return redirect(url_for("admin.dashboard"))

    error = None
    if request.method == "POST":
        ip = _client_ip()
        if not admin_configured():
            error = "Admin-Zugang ist serverseitig nicht konfiguriert."
        elif _is_locked(ip):
            error = "Zu viele Fehlversuche. Bitte in einigen Minuten erneut versuchen."
        elif not _check_csrf():
            error = "Sitzung abgelaufen. Bitte lade die Seite neu."
        elif _check_credentials(request.form.get("username"), request.form.get("password")):
            _reset_failures(ip)
            # Session-Fixation vermeiden: alte Session-Inhalte verwerfen.
            session.clear()
            session["admin_authenticated"] = True
            session.permanent = True
            return redirect(_safe_next(request.args.get("next")))
        else:
            _register_failure(ip)
            error = "Benutzername oder Passwort ist falsch."

    resp = Response(render_template("admin/login.html", error=error, csrf_token=csrf_token()))
    resp.headers["X-Robots-Tag"] = "noindex, nofollow"
    return resp


@admin_bp.route("/logout", methods=["POST"])
def logout():
    if not _check_csrf():
        abort(400)
    session.clear()
    return redirect(url_for("admin.login"))


@admin_bp.route("/")
@login_required
def dashboard():
    newsletter = db.get_newsletter_signups()
    interest = db.get_interest_submissions()
    for row in interest:
        row["config_display"] = format_config(row["config"])
    try:
        analytics = db.get_analytics()
    except Exception:
        analytics = None
    resp = Response(
        render_template(
            "admin/dashboard.html",
            newsletter=newsletter,
            interest=interest,
            analytics=analytics,
            csrf_token=csrf_token(),
        )
    )
    resp.headers["X-Robots-Tag"] = "noindex, nofollow"
    return resp


# --- Loeschen ---------------------------------------------------------------
@admin_bp.route("/newsletter/<int:signup_id>/delete", methods=["POST"])
@login_required
def delete_newsletter(signup_id):
    if not _check_csrf():
        abort(400)
    db.delete_newsletter(signup_id)
    return redirect(url_for("admin.dashboard"))


@admin_bp.route("/interest/<int:submission_id>/delete", methods=["POST"])
@login_required
def delete_interest(submission_id):
    if not _check_csrf():
        abort(400)
    db.delete_interest(submission_id)
    return redirect(url_for("admin.dashboard"))


# --- Excel-Export -----------------------------------------------------------
_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _safe_cell(value):
    """Schutz gegen Excel/CSV-Formel-Injection.

    Zellen, die mit =, +, - oder @ beginnen, werden mit einem fuehrenden
    Apostroph entschaerft, damit Excel sie nicht als Formel ausfuehrt.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def _fmt_dt(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value or ""


_GROUP_LABELS = None


def _group_labels():
    """Gruppen-IDs (neck, pickups, ...) auf lesbare Labels (Hals, Pick-Ups, ...).

    Lazy-Import von app.OPTIONS, um Zirkularimport beim Modulladen zu vermeiden.
    """
    global _GROUP_LABELS
    if _GROUP_LABELS is None:
        try:
            from app import OPTIONS
            _GROUP_LABELS = {k: v.get("label", k) for k, v in OPTIONS.items()}
        except Exception:
            _GROUP_LABELS = {}
    return _GROUP_LABELS


def format_config(config):
    """Konfig-dict -> lesbare Liste [(Label, Wert)].

    Gespeichert wird pro Gruppe {"id": ..., "name": ...}; wir zeigen den Namen.
    """
    if not config or not isinstance(config, dict):
        return []
    labels = _group_labels()
    out = []
    for key, val in config.items():
        label = labels.get(key, key)
        if isinstance(val, dict):
            name = val.get("name") or val.get("id") or ""
        else:
            name = str(val)
        out.append((label, name))
    return out


def _fmt_config(config):
    """Einzeilige, lesbare Konfiguration fuer den Excel-Export."""
    pairs = format_config(config)
    return "; ".join(f"{label}: {value}" for label, value in pairs)


@admin_bp.route("/export.xlsx")
@login_required
def export_excel():
    # Import hier, damit die App auch ohne openpyxl startet (nur Export braucht es).
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Newsletter"
    ws1.append(["ID", "E-Mail", "Einwilligung", "Angemeldet am"])
    for row in db.get_newsletter_signups():
        ws1.append([
            row["id"],
            _safe_cell(row["email"]),
            "ja" if row["consent"] else "nein",
            _fmt_dt(row["created_at"]),
        ])

    ws2 = wb.create_sheet("Interessenten")
    ws2.append([
        "ID", "Art", "E-Mail", "Name", "Straße", "PLZ", "Stadt",
        "Land", "Beitrag (€)", "Konfiguration", "Eingetragen am",
    ])
    for row in db.get_interest_submissions():
        ws2.append([
            row["id"],
            _safe_cell(row["kind"]),
            _safe_cell(row["email"]),
            _safe_cell(row["name"]),
            _safe_cell(row["street"]),
            _safe_cell(row["zip"]),
            _safe_cell(row["city"]),
            _safe_cell(row["country"]),
            row["contribution"],
            _safe_cell(_fmt_config(row["config"])),
            _fmt_dt(row["created_at"]),
        ])

    # Kopfzeilen fett
    for ws in (ws1, ws2):
        for cell in ws[1]:
            cell.font = Font(bold=True)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"bass-emails_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    resp = send_file(
        buffer,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )
    resp.headers["X-Robots-Tag"] = "noindex, nofollow"
    return resp
